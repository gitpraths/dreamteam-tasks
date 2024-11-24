# Reading Excel Files

import openpyxl

# Load workbook and select a sheet
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']

# Access a cell's value
print(sheet['A1'].value)

# Iterate over rows
for row in sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3):
    for cell in row:
        print(cell.value)
# Use load_workbook() to open an existing Excel file.
# Access individual sheets by name.
# Access cell values using square brackets or loops.

# Writing in Excel
        
# Modify an existing sheet
sheet['B2'] = 'Updated Value'

# Save changes
wb.save('updated_example.xlsx')

# Modify cells directly by assigning new values.
# Save your changes using the save() method.

# Creating a New Workbook
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
new_sheet.title = 'MySheet'

# Write data
new_sheet['A1'] = 'Hello'
new_sheet['A2'] = 'World'

new_wb.save('new_file.xlsx')

# Create a new workbook with Workbook().
# Set sheet titles and write data to cells.

# Applying Styles
from openpyxl.styles import Font, Alignment

bold_font = Font(bold=True)
center_align = Alignment(horizontal='center')

sheet['A1'].font = bold_font
sheet['A1'].alignment = center_align

wb.save('styled_example.xlsx')

# Use Font and Alignment for formatting.
# Apply styles to cells by setting properties like font and alignment.

# Adding Charts
from openpyxl.chart import BarChart, Reference
# Insert data
data = [
    ['Month', 'Sales'],
    ['Jan', 500],
    ['Feb', 700],
    ['Mar', 300],
]
for row in data:
    sheet.append(row)

# Create a chart
chart = BarChart()
values = Reference(sheet, min_col=2, min_row=2, max_row=4)
chart.add_data(values, titles_from_data=True)
sheet.add_chart(chart, 'E5')

wb.save('chart_example.xlsx')

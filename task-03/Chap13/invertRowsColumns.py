import openpyxl
import sys

def invert_rows_and_columns(filename):
 
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
       
            ws.cell(row=col, column=row, value=ws.cell(row=row, column=col).value)

    wb.save(filename)
    print(f"Spreadsheet rows and columns have been inverted in '{filename}'")

def main():

    if len(sys.argv) != 2:
        print("Usage: invertRowsColumns.py FILENAME")
        sys.exit(1)

    filename = sys.argv[1]

    try:
        invert_rows_and_columns(filename)
    except openpyxl.utils.exceptions.InvalidFileException:
        print(f"Error: The file '{filename}' is not a valid Excel file.")
        sys.exit(1)
    except Exception as e:
        print(f"An error occurred: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
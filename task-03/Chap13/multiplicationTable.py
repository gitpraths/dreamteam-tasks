import sys
import openpyxl

def create_multiplication_table(table_size):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    
    column_width = max(table_size**2, 12)
    worksheet.column_dimensions['A'].width = column_width
    
    for column_index in range(1, table_size + 1):
        worksheet.cell(row=1, column=column_index + 1, value=column_index)
        worksheet.cell(row=column_index + 1, column=1, value=column_index)
    
    for row_index in range(1, table_size + 1):
        for column_index in range(1, table_size + 1):
            product = row_index * column_index
            worksheet.cell(row=row_index + 1, column=column_index + 1, value=product)
    
    workbook.save('multiplication_table.xlsx')
    print('Multiplication table has been successfully created.')

def main():
    if len(sys.argv) != 2:
        print("Usage: python multiplicationTable.py TABLE_SIZE", file=sys.stderr)
        sys.exit(1)
    
    try:
        table_size = int(sys.argv[1])
        if table_size <= 0:
            print("Please enter a positive integer for the table size.", file=sys.stderr)
            sys.exit(1)
        
        create_multiplication_table(table_size)
        
    except ValueError:
        print("Error: The provided argument is not a valid integer for the table size.", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()
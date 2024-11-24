# Loading Workbooks:

import openpyxl  
wb = openpyxl.load_workbook('example.xlsx')  
sheet = wb.active

# Accessing Cells:
print(sheet['A1'].value)  # Access cell A1  
print(sheet.cell(row=1, column=1).value)  # Using indices
#Reading and Writing Data

# Reading Data:
for row in sheet.iter_rows(min_row=1, max_row=5):
    for cell in row:
        print(cell.value)

# Writing Data:
sheet['B1'] = "Hello, World!"
wb.save('updated_file.xlsx')

# Formulas: Add directly:

sheet['C1'] = "=SUM(A1:A10)"

# Advanced Features
# Styling Cells:
from openpyxl.styles import Font  
sheet['A1'].font = Font(size=14, bold=True)

# Merging and Freezing:
sheet.merge_cells('A1:B1')  
sheet.freeze_panes = 'A2'

# Creating Charts:
from openpyxl.chart import BarChart, Reference  
chart = BarChart()  
data = Reference(sheet, min_col=2, min_row=1, max_row=10)  
chart.add_data(data, titles_from_data=True)  
sheet.add_chart(chart, 'E5')
# Alternatives to Excel
# Tools like LibreOffice Calc and OpenOffice Calc support .xlsx files and can also be used for spreadsheet operations.

# Conclusion

# The openpyxl module enables Python to automate and streamline spreadsheet tasks, including reading, writing, styling, and generating charts. It is a powerful tool for data management and analysis.
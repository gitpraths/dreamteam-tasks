import sys
import openpyxl

def export_spreadsheet_to_text_files(workbook_filename):
    workbook = openpyxl.load_workbook(workbook_filename)
    worksheet = workbook.active
    
    max_column = worksheet.max_column
    
    for column_index in range(1, max_column + 1):
        file_path = f'column{column_index}.txt'
        with open(file_path, 'w', encoding='utf-8') as file:
            for row_index in range(1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row_index, column=column_index).value
                if cell_value is not None:
                    file.write(str(cell_value))
                    if row_index < worksheet.max_row:
                        file.write('\n')

def main():
    if len(sys.argv) != 2:
        print("Usage: python spreadsheetToTextFiles.py WORKBOOK_FILENAME", file=sys.stderr)
        sys.exit(1)
    
    workbook_filename = sys.argv[1]
    
    try:
        export_spreadsheet_to_text_files(workbook_filename)
        print('Spreadsheet contents have been successfully written to text files.')
    except FileNotFoundError:
        print(f"Error: The file '{workbook_filename}' does not exist.", file=sys.stderr)
        sys.exit(1)
    except openpyxl.utils.exceptions.InvalidFileException:
        print(f"Error: The file '{workbook_filename}' is not a valid Excel file.", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()